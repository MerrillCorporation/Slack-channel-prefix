import os
import logging
import logging.config
import time
import yaml

from openpyxl import load_workbook
from copy import copy
from slack_bolt import Ack, App
from slack_bolt.adapter.socket_mode import SocketModeHandler
from slack_sdk.errors import SlackApiError

# Function to load yaml logging configuration
def logging_config(
    default_path="logging.yaml", default_level=logging.INFO, env_key="LOG_CFG"
):
    """
    Setup logging configuration
    """
    path = default_path
    value = os.getenv(env_key, None)
    if value:
        path = value
    if os.path.exists(path):
        with open(path, "rt") as f:
            config = yaml.safe_load(f.read())
        logging.config.dictConfig(config)
    else:
        logging.basicConfig(level=default_level)

logging_config()

# Create Logging Object
LOG = logging.getLogger(__name__)

# Token Authorization
try:
    LOG.info("Getting slack_bot token")
    botToken = os.environ.get("SLACK_BOT_TOKEN")
    userToken = os.environ.get("SLACK_USER_TOKEN")

    # Initializes app with bot token and socket mode handler
    app = App(token=botToken)
    appUser = App(token=userToken)
except:
    LOG.error("invalid/no token")
    
try:
    LOG.info("Getting slack_user token")
    usrToken = os.environ.get("SLACK_USER_TOKEN")
    userApp = App(token=usrToken)
except:
    LOG.error("Could not retrieve user token")

# Common header use
headers = {"Accept": "application/json", "Content-Type": "application/json"}


def main_function():
    LOG.info('Begin main function')
    channel_info_dict = get_channel_info()
    ch_info_to_excel(channel_info_dict)
    pre_fix_notify(channel_info_dict)

def ch_info_to_excel(channel_id_dict):
    try:
        # Run local excel file, saved in location
        LOG.info("Opening Excel Workbook")
        wb_name = "prefix_sheet.xlsx"
        wb = load_workbook(wb_name)
        sheet_name = "Sheet 1"
        sheet = wb[sheet_name]
    except:
        LOG.error("Unable to open workbook")

    try: 
        LOG.info("Iterate through excel sheet, write in channel id, name, and creator id")

        # Begin at second row of sheet
        i = 2
        for ch_id in channel_id_dict:
            sheet.cell(row=i, column=1, value=ch_id)    
            sheet.cell(row=i, column=2, value=channel_id_dict[ch_id][0]) 
            sheet.cell(row=i, column=3, value=channel_id_dict[ch_id][1]) 
            i += 1     
        wb.save(filename=wb_name) 
    except:
        LOG.error('Could not iterate and write data to excel')     


def get_channel_info():
    # Read in text file for channel names
    channel_name_list = []
    try:
        LOG.info('Open and reading text file')
        with open('channel_names.txt') as f:
            for line in f:
                channel_name_list.append(line.rstrip())
    except:
        LOG.info('Could note read from text file')

    # Initial request to grab slack channels 
    try:   
        LOG.info('Retrieving first page of slack channels')
        result = app.client.conversations_list(limit=150)

    except:
        LOG.error('Could not retrieve first page of channels')

    # Initial cursor (First page of channels)
    next_cursor = result['response_metadata']['next_cursor']

    num_page = 0
    num_ch = len(result['channels'])
    channel_info_dict = {}

    # Iterate through all public slack channels 
    while True:
        # Iterate through slack channels for matches
        for channel_name in channel_name_list:  
            for channel in result['channels']:
                ch_name = channel_name.casefold()
                result_ch_name = channel['name'].casefold()
                if ch_name == result_ch_name:
                    channel_info_dict[channel['id']] = (channel_name, channel['creator'])  
        LOG.info(f'Current page: {num_page}\nChannels Checked: {num_ch}')

        # Checks if last page of channels ->  exit loop. < 149 due to each request for 150 channels.  
        if len(result['channels']) < 149:
            break

        try:
            LOG.info('Retrieving next set of channels')
            result = app.client.conversations_list(limit=150, cursor=next_cursor)

        except SlackApiError as e:
            if e.response.status_code == 429:
                delay = int(e.response.headers['Retry-After'])
                LOG.error(f'RATE LIMIT EXCEEDED, RETRYING IN {delay} SECONDS')
                time.sleep(delay)
                continue 
            else:
                LOG.error('Could not retrieve channel list')

        # Calculates page number and number of channels checked                    
        num_page += 1
        num_ch += len(result['channels'])

        # Retrieves cursor for the next page
        next_cursor = result['response_metadata']['next_cursor']    

    return channel_info_dict


def pre_fix_notify(channel_info_dict):
    LOG.info("Begin sending notifications")
    # Gets channel information
    for channel in channel_info_dict:
        ch_name = channel_info_dict[channel][0]
        creator_id = channel_info_dict[channel][1]

        # Send message to channel creator
        try: 
            LOG.info("Posting message")
            text = f"Hi <@{creator_id}>!\nA prefix will be added to your channel: #{ch_name}\nOn: **/**/****\nThis will be done to help organize our slack workspace. If you have any questions or concerns, please contact @squad_productivity_eng :partyblob:"
            app.client.chat_postMessage(
                channel=creator_id,
                blocks=[
                    {
                        "type": "section",
                        "text": {
                            "type": "mrkdwn",
                            "text": text
                        },
                    }
                ],
                text=text,
            )
        except:
            LOG.error('Could not post message to channel creator U02AT40NEF4')

# Run Main Function
main_function()


@app.event("message")
def handle_event_msg():
    return

# Start your app
if __name__ == "__main__":
    LOG.info("Starting app")
    SocketModeHandler(app, os.environ["SLACK_APP_TOKEN"]).start()
